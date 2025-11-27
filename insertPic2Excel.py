import os
import re
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from typing import Optional, List, Tuple
import fitz # PyMuPDF 导入时使用 fitz
from PIL import Image as PILImage
from difflib import SequenceMatcher
from datetime import datetime # 引入 datetime 库获取当前日期
from openpyxl.utils.cell import coordinate_from_string
import openpyxl.utils.units as units
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

TARGET_MAX_WIDTH = 400  # 目标最大宽度（像素）
TARGET_MAX_HEIGHT = 400 # 目标最大高度（像素）
# openpyxl 默认行高单位 (1/400英寸)，像素转行高的近似比例，需要根据字体调整
ROW_HEIGHT_SCALE = 0.75 # 假设 1 像素大约等于 0.75 openpyxl 高度单位

def get_max_image_row(sheet) -> int:
    """
    遍历工作表中的图片，找到它们覆盖到的最大行号，兼容 OneCellAnchor 和 TwoCellAnchor。
    """
    max_image_end_row = 0
    
    # openpyxl 图像对象列表位于内部属性 _images 或 _charts
    # 遍历所有 Drawing 对象（包括图片、图表等）以覆盖所有情况
    for drawing in sheet._images + sheet._charts:
        anchor = drawing.anchor
        
        # ----------------------------------------------------
        # 情况 A: TwoCellAnchor (有 from 和 to 属性)
        # ----------------------------------------------------
        if isinstance(anchor, TwoCellAnchor):
            # TwoCellAnchor 的 .to.row 是结束行索引（从 0 开始）
            image_end_row = anchor.to.row + 1  # 索引转 Excel 行号 (1-based)
        
        # ----------------------------------------------------
        # 情况 B: OneCellAnchor (只有 from 属性，需要计算高度)
        # ----------------------------------------------------
        elif isinstance(anchor, OneCellAnchor):
            # 1. 获取起始行 (from.row 是起始行索引)
            start_row = anchor._from.row + 1 # 索引转 Excel 行号
            
            # 2. 获取图片高度 (图片在 Excel中的尺寸，单位 EMU)
            if hasattr(drawing, 'height') and drawing.height is not None:
                image_height_emu = drawing.height
            elif hasattr(drawing.drawing, 'height') and drawing.drawing.height is not None:
                image_height_emu = drawing.drawing.height
            else:
                # 无法获取高度，只能假设它只占一行
                max_image_end_row = max(max_image_end_row, start_row)
                continue

            # 3. 计算图片占用的行数（这是一个复杂的近似计算）
            # 遍历起始行及之后的行，累加其高度，直到覆盖图片高度
            current_height_emu = 0
            end_row = start_row
            
            # 迭代计算图片会覆盖到的行
            while current_height_emu < image_height_emu:
                # 检查行维度是否存在，如果不存在，使用默认高度
                if end_row in sheet.row_dimensions and sheet.row_dimensions[end_row].customHeight:
                    row_height_pts = sheet.row_dimensions[end_row].height
                    row_height_emu = units.points_to_pixels(row_height_pts) * 9525 # 粗略转换
                else:
                    # 使用默认行高 (约 15 points = 190500 EMU)
                    row_height_emu = 190500 

                current_height_emu += row_height_emu
                
                # 如果这是第一行，且高度足够，不递增行号，否则递增
                if end_row == start_row and current_height_emu >= image_height_emu:
                    # 图片的高度小于或等于第一行，结束行就是起始行
                    break 
                elif end_row > start_row and current_height_emu >= image_height_emu:
                    # 图片覆盖了前几行，但最后一行只需要一部分空间
                    break
                
                end_row += 1
            
            image_end_row = end_row
            
        else:
            # 遇到其他未知的锚点类型，跳过
            continue

        # 更新全局最大行号
        max_image_end_row = max(max_image_end_row, image_end_row)
        
    return max_image_end_row

def resize_image_for_excel(path_to_insert: str) -> Optional[ExcelImage]:
    """
    加载图片文件，将其尺寸等比例缩放至 Excel 目标尺寸，并返回 openpyxl Image 对象。

    参数:
    path_to_insert (str): 图片文件的完整路径。

    返回:
    Optional[ExcelImage]: 调整好尺寸的 openpyxl Image 对象，如果失败则返回 None。
    """
    try:
        # 1. 加载图片以获取原始尺寸
        img = ExcelImage(path_to_insert)
        
        original_width = img.width
        original_height = img.height
        
        # 2. 计算缩放比例
        width_ratio = TARGET_MAX_WIDTH / original_width
        height_ratio = TARGET_MAX_HEIGHT / original_height
        
        # 取较小的比例进行等比例缩放，并确保图片不会被放大
        scale_ratio = min(width_ratio, height_ratio, 1.0) 
        
        new_width = int(original_width * scale_ratio)
        new_height = int(original_height * scale_ratio)
        
        # 3. 应用新尺寸
        img.width = new_width
        img.height = new_height
        
        return img
    except Exception as e:
        # 如果是图片格式错误或其他加载问题，在这里捕获
        print(f"缩放图片 {os.path.basename(path_to_insert)} 时发生错误: {e}")
        return None

def adjust_row_height(sheet, current_row: int, image_height_px: int) -> None:
    """
    根据图片在 Excel 中的预期高度（像素），调整工作表中对应行的行高。
    """
    # openpyxl 高度单位与像素的近似转换
    required_height = image_height_px * ROW_HEIGHT_SCALE
    
    # 确保新行高不会小于图片所需的最小高度
    current_row_dim = sheet.row_dimensions[current_row]
    if current_row_dim.height is None or required_height > current_row_dim.height:
         current_row_dim.height = required_height
    # 否则保持现有行高（如果更高）


def find_first_empty_row_and_format(file_path: str) -> Optional[Tuple[str, str]]:
    """
    找出一个 Excel 文件中，名称相似的工作表内没有内容的第一行。
    将该行的 A 到 N 列变成绿色，并返回工作簿名称和下一行的起始单元格位置。

    参数:
    file_path (str): XLSX 文件的完整路径。

    返回:
    Optional[Tuple[str, str]]: (工作簿名称, 下一行的起始单元格位置)，如果找到并格式化成功；
                               如果文件不存在或找不到目标工作表，则返回 None。
    """
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在 -> {file_path}")
        return None

    # 定义目标工作表名称的关键词（不区分大小写）
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在 -> {file_path}")
        return None

    # --- 1. 获取当前月份标记 ---
    current_month_marker = datetime.now().strftime("%Y-%m") # 格式如 '2025-11'
    print(f"当前月份标记: {current_month_marker}")
    
    # --- 2. 标记列和格式化列定义 ---
    MARKER_COLUMN_LETTER = 'O' # 使用 O 列作为月份标记列
       
    # 定义关键词和权重
    RECEIPT_KEYWORDS = ["Receipts", "收据", "支出收据", "Invoice"]
    TARGET_YEAR = "2025"
    EXCLUDE_YEAR = "2024" # 新增排除关键词

    # 权重设定
    SCORE_TARGET_YEAR = 100        # 包含 2025 给予高分
    PENALTY_EXCLUDE_YEAR = -200    # 包含 2024 给予极低的分数 (确保其被忽略)
    SCORE_RECEIPT_BASE = 10
    
    # 定义绿色填充样式
    # RGB 颜色代码，例如 '00FF00' 是纯绿色，'C6EFCE' 是 Excel 中的浅绿色
    GREEN_FILL = PatternFill(start_color='006100', end_color='006100', fill_type='solid')
       
    # A 列到 N 列的列名
    COLUMNS_TO_FORMAT = [chr(ord('A') + i) for i in range(14)] # A, B, C, ..., N

    try:
        # 1. 加载工作簿
        workbook = load_workbook(file_path)
    except Exception as e:
        print(f"错误: 无法加载工作簿 -> {e}")
        return None
    
    target_sheet = None
    
    # 2. 遍历工作表，寻找目标名称
    best_sheet_name = None
    max_score = -999 # 初始化为更小的负数
    
    print("开始评分工作表:")

    for sheet_name in workbook.sheetnames:
        lower_name = sheet_name.lower()
        current_score = 0
        
        # 1.1. 评分：惩罚 2024
        if EXCLUDE_YEAR in lower_name:
            current_score += PENALTY_EXCLUDE_YEAR # 极低的惩罚分
        
        # 1.2. 评分：目标年份 2025 (最高优先级)
        if TARGET_YEAR in lower_name:
            current_score += SCORE_TARGET_YEAR
            
        # 1.3. 评分：收据/Receipts 关键词
        max_receipt_ratio = 0.0
        
        for keyword in RECEIPT_KEYWORDS:
            # 计算工作表名和核心关键词的最高相似度
            ratio = SequenceMatcher(None, lower_name, keyword.lower()).ratio()
            max_receipt_ratio = max(max_receipt_ratio, ratio)
        
        # 将相似度百分比作为权重添加到基础分数上
        current_score += SCORE_RECEIPT_BASE * max_receipt_ratio
        
        print(f"  - '{sheet_name}' 评分: {current_score:.2f} (2025: {TARGET_YEAR in lower_name}, 2024: {EXCLUDE_YEAR in lower_name})")

        # 1.4. 选择最高分
        if current_score > max_score:
            max_score = current_score
            best_sheet_name = sheet_name
        
    if best_sheet_name is None or max_score <= PENALTY_EXCLUDE_YEAR: # 确保得分极低的表不被选中
        print("未找到任何相关的工作表。")
        return None

    target_sheet = workbook[best_sheet_name]
    print(f"✅ 最终选定目标工作表: {best_sheet_name} (最高分: {max_score:.2f})")

    # 3. 查找第一个空行
    # max_row 是包含内容的行数（或至少有格式的行数）。
    # 我们可以从 max_row + 1 开始检查，或者从第一行开始检查直到遇到空行。
    
    first_untagged_empty_row = None
    last_content_row = 0
    max_row_to_check = target_sheet.max_row + 100
    # 循环从第 1 行开始，确保检查所有可能的行
    print(f"正在扫描工作表 '{best_sheet_name}' (最大行号 {target_sheet.max_row}) 寻找最后的内容行...")
    for row_num in range(1,max_row_to_check):
        
        is_row_content_present = False
    
        # 只检查 A-N 列是否有内容
        for col_letter in COLUMNS_TO_FORMAT:
            cell = target_sheet[f'{col_letter}{row_num}']
            
            # 排除已标记行中的空值，以免将标记行算作内容行
            marker_cell = target_sheet[f'{MARKER_COLUMN_LETTER}{row_num}'] # O 列标记

            # 检查 A-N 列是否有内容，同时忽略已被当前月份标记的行（如果标记行A-N是空的）
            if cell.value is not None and str(cell.value).strip() != "":
                # 找到内容了，更新最大行号，并退出内层循环
                last_content_row = row_num
                is_row_content_present = True
                break
                
            # 如果行是空的，但 O 列已经被标记了，我们仍然认为它已经被处理过了
            if col_letter == COLUMNS_TO_FORMAT[0] and marker_cell.value is not None and str(marker_cell.value).strip() == current_month_marker:
                # 标记行，即使A-N是空的，也应该算作“已处理”
                last_content_row = row_num
                # 不需要 break，让它继续检查 A-N 列
                
        # 优化：如果当前行和之前的行是空的，并且超过了上一次找到的最大内容行，则可以提前退出扫描。
        # 但为了稳妥，我们让它继续扫描到 max_row + 100 的边界。
    if target_sheet._images:
        max_image_row = get_max_image_row(target_sheet)
    else:
        max_image_row = 0
    
    last_known_row = max(last_content_row, max_image_row)
    print(f"✅ 找到最后有内容的行 (或已标记行): 第 {last_known_row} 行")

    # -------------------------------------------------------------------
    # 4. 确定下一个插入行 (last_content_row 的下一行)
    # -------------------------------------------------------------------
    first_untagged_empty_row = last_known_row + 1 

    # -------------------------------------------------------------------
    # 5. 格式化该行 (新逻辑：只格式化新确定的行)
    # -------------------------------------------------------------------

    # ⚠️ 修复：现在需要检查新确定的行是否已经被标记过！
    marker_cell_next = target_sheet[f'{MARKER_COLUMN_LETTER}{first_untagged_empty_row}']
    if marker_cell_next.value is not None and str(marker_cell_next.value).strip() == current_month_marker:
        # 这种情况理论上不应该发生，除非 Excel 尾部有重复的标记行。
        print(f"警告: 第 {first_untagged_empty_row} 行已被标记。跳过格式化。")
        # 如果该行已标记，则下一行才是真正的空行
        first_untagged_empty_row += 1 
        
        # 重新获取单元格
        marker_cell_to_set = target_sheet[f'{MARKER_COLUMN_LETTER}{first_untagged_empty_row}']
    else:
        marker_cell_to_set = marker_cell_next

    # 格式化 A-N 列
    for col_letter in COLUMNS_TO_FORMAT:
        cell = target_sheet[f'{col_letter}{first_untagged_empty_row}']
        cell.fill = GREEN_FILL
        
    # 设置月份标记 (关键步骤)
    marker_cell_to_set.value = current_month_marker
        
    print(f"✅ 已将第 {first_untagged_empty_row} 行设置为绿色，并标记 '{current_month_marker}' 在 {MARKER_COLUMN_LETTER} 列。")
        
    # 5. 保存工作簿
    try:
        workbook.save(file_path)
        print("✅ 工作簿已保存。")
    except Exception as e:
        print(f"错误: 无法保存工作簿。请确保文件未被其他程序打开。-> {e}")
        return None
    
    # 6. 返回结果
    workbook_name = best_sheet_name
    # 下一行（数据输入行）的起始位置
    next_row_start_cell = f'A{first_untagged_empty_row + 1}'
    
    return (workbook_name, next_row_start_cell)

def insert_images_to_excel_with_pdf(excel_path: str, file_paths: List[str] ):
    """
    将多个文件 (PNG, JPG, JPEG, PDF) 作为图片插入到指定的 Excel 工作表中。
    PDF 文件会被转换为临时 PNG 图像后插入。

    参数:
    excel_path (str): 要创建或修改的 Excel 文件路径。
    file_paths (List[str]): 文件（图片或 PDF）的完整路径列表。
    """

    # 允许的图片文件扩展名
    ALLOWED_EXTENSIONS = ('.png', '.jpg', '.jpeg')
    
    # --- 1. 准备 Excel 工作簿 ---
    # ... (这部分与原代码保持一致，用于加载/创建工作簿和工作表)
    
    # 检查文件是否存在，如果存在则加载，否则创建新的
    if os.path.exists(excel_path):
        try:
            workbook = load_workbook(excel_path)
        except Exception as e:
            print(f"警告: 无法加载现有文件 {excel_path}，将创建新的工作簿。错误: {e}")
            workbook = Workbook()
    else:
        workbook = Workbook()
        
    print(f"正在分析 Excel 文件 '{excel_path}'，寻找插入位置...")
    
    # ⚠️ 确保 find_first_empty_row_and_format 返回 (best_sheet_name, next_row_start_cell)
    result = find_first_empty_row_and_format(excel_path)
    
    if result is None:
        print("❌ 无法插入图片: find_first_empty_row_and_format 无法找到有效的插入位置。")
        return # 提前退出
        
    # 覆盖 sheet_name 和 start_cell 变量
    actual_sheet_name, actual_start_cell = result        
    
    try:
        # 重新加载工作簿，包含 find_first_empty_row_and_format 写入的绿色标记行
        workbook = load_workbook(excel_path)
    except Exception as e:
        print(f"错误: 无法重新加载工作簿 {excel_path}。错误: {e}")
        return
    
    sheet = workbook[actual_sheet_name]
    
    # --- 2. 确定图片起始位置 ---

    col_str, row_idx = coordinate_from_string(actual_start_cell) # 使用确定的单元格
    
    current_row = row_idx
    success_count = 0

    print(f"开始向工作表 '{actual_sheet_name}' 插入文件...")

    # --- 3. 遍历文件并插入图片 ---
    for i, file_path in enumerate(file_paths):
        if not os.path.exists(file_path):
            print(f"跳过: 文件不存在 - {file_path}")
            continue

        file_ext = os.path.splitext(file_path)[1].lower()
        file_name = os.path.basename(file_path)
        
        # 默认插入路径为原始文件路径
        path_to_insert = file_path 
        pdf_dir = os.path.dirname(file_path)
        # 🎯 新的持久临时目录
        pdf_img_dir = os.path.join(pdf_dir, "_temp_img") 
        os.makedirs(pdf_img_dir, exist_ok=True) # 确保目录存在
        # --- PDF 特殊处理 ---
        if file_ext == '.pdf':
            try:
                # 尝试打开 PDF
                pdf_doc = fitz.open(file_path)                  
                # 遍历 PDF 的每一页
                for page_num in range(len(pdf_doc)):
                    page = pdf_doc.load_page(page_num)
                    
                    # 设置渲染参数 (dpi可以控制图片清晰度)
                    matrix = fitz.Matrix(200 / 72, 200 / 72) # 200 DPI
                    pix = page.get_pixmap(matrix=matrix, alpha=False)

                    # 生成临时 PNG 文件路径
                    temp_png_path = os.path.join(pdf_img_dir, f"{file_name}_page{page_num+1}.png")
                    
                    # 保存为 PNG
                    pix.save(temp_png_path)
                    
                    # 设置为当前插入路径，并更新文件名以在打印中显示页码
                    path_to_insert = temp_png_path
                    display_name = f"{file_name} (Page {page_num+1})"

                    # 执行插入操作 (将插入操作放入循环内，处理PDF的每一页)
                    try:
                        img = resize_image_for_excel(path_to_insert)
                        adjust_row_height(sheet, current_row, img.height)
                        anchor_cell = f"{col_str}{current_row}"
                        sheet.add_image(img, anchor_cell)
                        print(f"成功插入: '{display_name}'，位置: {anchor_cell}")
                        current_row += 1 
                        success_count += 1
                    except Exception as e:
                        print(f"插入 PDF 图像 '{display_name}' 时发生错误: {e}")
                
                pdf_doc.close()
                continue # 处理完 PDF 后，跳到下一个文件
                
            except Exception as e:
                print(f"处理 PDF 文件 '{file_name}' 时发生错误: {e}")
                continue
        
        # --- 正常图片文件处理 ---
        elif file_ext not in ALLOWED_EXTENSIONS:
            print(f"跳过: 文件 '{file_name}' 格式 ({file_ext}) 不支持。")
            continue
        
        # 插入 PNG/JPG/JPEG 文件
        try:
            img = resize_image_for_excel(file_path)
            adjust_row_height(sheet, current_row, img.height)
            anchor_cell = f"{col_str}{current_row}"
            sheet.add_image(img, anchor_cell)
            print(f"成功插入: '{file_name}'，位置: {anchor_cell}")
            current_row += 1 
            success_count += 1
        except Exception as e:
            print(f"插入图片 '{file_name}' 时发生错误: {e}")

    # --- 4. 保存工作簿 ---
    try:
        workbook.save(excel_path)
        print("\n🎉 任务完成！")
        print(f"文件已保存至: {excel_path}")
        print(f"总共成功插入 {success_count} 张图片/页面。")
    except Exception as e:
        print(f"\n致命错误: 无法保存 Excel 文件。请确保文件未被打开。错误: {e}")


# --- 使用示例 ---
if __name__ == "__main__":
    
    # ⚠️ 1. 替换为您的输出文件路径
    output_excel_file = os.path.join(os.getcwd(), "Image_PDF_Report.xlsx")
    
    # ⚠️ 2. 替换为您要插入的实际文件路径列表
    # 请确保这些文件路径在您的系统上是存在的！
    files_to_insert = [
        r"C:\Path\To\Your\Image1.png",
        r"C:\Path\To\Your\Document.pdf",    # 这是一个将被转换为多张图片的 PDF 文件
        r"C:\Path\To\Your\Photo2.jpg",
        r"C:\Path\To\Your\Another.pdf"
    ]
    

    # 3. 执行函数
    insert_images_to_excel_with_pdf(
        excel_path=output_excel_file,
        file_paths=files_to_insert,
    )